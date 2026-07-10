from django.shortcuts import render, redirect
from datetime import datetime, time, timedelta, date
from django.db import IntegrityError
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
from django.contrib.auth.hashers import check_password
from django.views.decorators.cache import never_cache

from .models import ExternalBooking
from openpyxl import load_workbook
import re
from .models import InternalPlayer
from django.db import transaction

from .models import InternalBooking

from .models import RegisterOverride


# =========================
# HOME PAGE
# =========================
@never_cache
def home(request):
    # Clear any active booking state when navigating to home
    keys_to_clear = [
        "selected_slot", "selected_date", "booking_id",
        "internal_selected_slot", "internal_selected_date", "internal_booking"
    ]
    for key in keys_to_clear:
        if key in request.session:
            del request.session[key]

    if request.method == "POST":
        user_type = request.POST.get("user_type")

        if user_type == "internal":
            return redirect("booking:internal_login")

        elif user_type == "external":
            return redirect("booking:external_slots")

    return render(request, "home.html")


# =========================
# EXTERNAL SLOT SELECTION
# =========================
@never_cache
def external_slots(request):
    date_str = request.GET.get("date")

    # default = tomorrow
    if date_str:
        selected_date = datetime.strptime(date_str, "%Y-%m-%d")
    else:
        selected_date = datetime.today() + timedelta(days=1)

    weekday = selected_date.weekday()

    # Slot rules
    if weekday <= 4:  # Monday–Friday
        start_hour = 8
        end_hour = 15
    else:  # Saturday–Sunday
        start_hour = 7
        end_hour = 21

    # Generate slots
    slots = []
    for hour in range(start_hour, end_hour):
        start = time(hour, 0).strftime("%I:%M %p")
        end = time(hour + 1, 0).strftime("%I:%M %p")
        slots.append(f"{start} - {end}")

    # Already booked slots
    booked_slots = ExternalBooking.objects.filter(
        date=selected_date.date()
    ).values_list("slot", flat=True)

    if request.method == "POST":
        request.session["selected_slot"] = request.POST.get("selected_slot")
        request.session["selected_date"] = selected_date.strftime("%Y-%m-%d")
        return redirect("booking:external_details")

    tomorrow = datetime.today() + timedelta(days=1)
    day_after_tomorrow = datetime.today() + timedelta(days=2)

    context = {
        "slots": slots,
        "booked_slots": booked_slots,
        "day": selected_date.strftime("%A"),
        "date": selected_date.strftime("%Y-%m-%d"),

        # only change here
        "tomorrow": tomorrow.strftime("%Y-%m-%d"),
        "day_after_tomorrow": day_after_tomorrow.strftime("%Y-%m-%d"),

        "selected_date": selected_date.strftime("%Y-%m-%d"),
    }

    return render(request, "external_slots.html", context)


# =========================
# EXTERNAL DETAILS
# =========================
@never_cache
def external_details(request):
    selected_slot = request.session.get("selected_slot")
    selected_date_str = request.session.get("selected_date")

    if not selected_slot or not selected_date_str:
        return redirect("booking:external_slots")

    if request.method == "POST":
        name = request.POST.get("name")
        mobile = request.POST.get("mobile")
        persons = request.POST.get("persons")
        receipt = request.FILES.get("receipt")

        selected_date = datetime.strptime(
            selected_date_str, "%Y-%m-%d"
        ).date()

        try:
            booking = ExternalBooking.objects.create(
                name=name,
                mobile=mobile,
                persons=persons,
                slot=selected_slot,
                date=selected_date,
                receipt=receipt,
            )

            request.session["booking_id"] = booking.id
            return redirect("booking:external_confirmation")

        except IntegrityError:
            return render(request, "external_details.html", {
                "error": "❌ This slot is already booked. Please select another slot."
            })

    return render(request, "external_details.html")


# =========================
# EXTERNAL CONFIRMATION
# =========================
@never_cache
def external_confirmation(request):
    booking_id = request.session.get("booking_id")

    if not booking_id:
        return redirect("booking:home")

    booking = ExternalBooking.objects.get(id=booking_id)

    context = {
        "day": booking.date.strftime("%A"),
        "date": booking.date.strftime("%Y-%m-%d"),
        "slot": booking.slot,
        "payment_status": booking.payment_status,
    }

    return render(request, "external_confirmation.html", context)


# ========================= 
# INTERNAL LOGIN 
# =========================
@never_cache
def internal_login(request):

    if request.method == "POST":

        email = request.POST.get("email")
        password = request.POST.get("password")

        student = Student.objects.filter(email=email).first()

        if student and check_password(password, student.password):

            request.session["internal_logged_in"] = True
            request.session["student_email"] = student.email
            request.session["student_username"] = student.username

            return redirect("booking:internal_slots")

        return render(request, "internal_login.html", {
            "error": "Invalid email or password"
        })

    return render(request, "internal_login.html")


# =========================
# INTERNAL SLOTS
# =========================
@never_cache
def internal_slots(request):

    # =========================
    # LOGIN PROTECTION
    # =========================
    if not request.session.get("internal_logged_in"):
        return redirect("booking:internal_login")

    # Logged-in student details
    student_email = request.session.get("student_email")
    student_username = request.session.get("student_username")

    # =========================
    # DATE HANDLING
    # =========================
    date_str = request.GET.get("date")

    today = date.today()
    tomorrow = today + timedelta(days=1)
    day_after_tomorrow = today + timedelta(days=2)

    if date_str:
        try:
            selected_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            selected_date = tomorrow
    else:
        selected_date = tomorrow

    weekday = selected_date.weekday()  # 0 = Monday

    # =========================
    # INTERNAL SLOT RULES
    # =========================
    slots = []

    if weekday <= 4:  # Monday – Friday

        slots = [
            "05:30 AM - 06:30 AM",
            "06:30 AM - 07:30 AM",
            "04:30 PM - 05:30 PM (Girls Only)",
            "05:30 PM - 06:30 PM",
            "06:30 PM - 07:30 PM",
            "07:30 PM - 08:30 PM",
            "08:30 PM - 09:30 PM",
        ]

    else:  # Saturday & Sunday

        for hour in range(7, 21):
            start = time(hour, 0).strftime("%I:%M %p")
            end = time(hour + 1, 0).strftime("%I:%M %p")
            slots.append(f"{start} - {end}")

    # =========================
    # BOOKED SLOT CHECK
    # =========================
    internal_booked = InternalBooking.objects.filter(
        date=selected_date
    ).values_list("slot", flat=True)

    booked_slots = set(internal_booked)

    # External bookings block only on weekends
    if weekday >= 5:

        external_booked = ExternalBooking.objects.filter(
            date=selected_date
        ).values_list("slot", flat=True)

        booked_slots.update(external_booked)

    # =========================
    # SLOT SELECTION
    # =========================
    if request.method == "POST":

        selected_slot = request.POST.get("selected_slot")

        if not selected_slot:
            return render(request, "internal_slots.html", {
                "slots": slots,
                "booked_slots": booked_slots,
                "tomorrow": tomorrow.strftime("%Y-%m-%d"),
                "day_after_tomorrow": day_after_tomorrow.strftime("%Y-%m-%d"),
                "selected_date": selected_date.strftime("%Y-%m-%d"),
                "day": selected_date.strftime("%A"),
                "date": selected_date.strftime("%Y-%m-%d"),
                "student_email": student_email,
                "student_username": student_username,
                "error": "Please select a slot."
            })

        request.session["internal_selected_slot"] = selected_slot
        request.session["internal_selected_date"] = selected_date.strftime("%Y-%m-%d")

        return redirect("booking:internal_details")

    # =========================
    # CONTEXT
    # =========================
    context = {

        "slots": slots,
        "booked_slots": booked_slots,

        "tomorrow": tomorrow.strftime("%Y-%m-%d"),
        "day_after_tomorrow": day_after_tomorrow.strftime("%Y-%m-%d"),

        "selected_date": selected_date.strftime("%Y-%m-%d"),
        "day": selected_date.strftime("%A"),
        "date": selected_date.strftime("%Y-%m-%d"),

        # Logged-in Student
        "student_email": student_email,
        "student_username": student_username,
    }

    return render(request, "internal_slots.html", context)


@never_cache
def internal_details(request):
    if not request.session.get("internal_logged_in"):
        return redirect("booking:internal_login")

    selected_date_str = request.session.get("internal_selected_date")
    selected_slot = request.session.get("internal_selected_slot")

    if not selected_date_str or not selected_slot:
        return redirect("booking:internal_slots")

    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        mobile = request.POST.get("mobile", "").strip()
        persons = int(request.POST.get("persons", 0))
        excel_file = request.FILES.get("excel_file")

        # 🔴 Rule 1: Minimum 8 persons
        if persons < 8:
            return render(request, "internal_details.html", {
                "error": "Minimum 8 persons are required for internal booking."
            })

        # 🔴 Slot & Date safety check
        selected_date = datetime.strptime(selected_date_str, "%Y-%m-%d").date()

        try:
            wb = load_workbook(excel_file)
            sheet = wb.active

            rows = list(sheet.iter_rows(min_row=2, values_only=True))
            valid_rows = [r for r in rows if r and r[0] and r[1]]

            # 🔴 Rule 2: Max 15 players
            if len(valid_rows) > 15:
                raise ValueError("Excel file must not contain more than 15 players.")

            # 🔴 Rule 3: Count mismatch
            if len(valid_rows) != persons:
                raise ValueError("Number of persons does not match Excel entries.")

            week_start = selected_date - timedelta(days=selected_date.weekday())
            week_end = week_start + timedelta(days=6)

            players_data = []

            for person_name, reg_no in valid_rows:
                reg_no = str(reg_no).strip()

                # 🔴 Rule 4: Register format
                if not re.fullmatch(r"(21|22|23|24|25)\d{14}", reg_no):
                    raise ValueError(
                        "Register number must start with 21–25 and contain exactly 16 digits."
                    )

                # 🔴 Rule 5: Weekly limit (2 per week)
                weekly_count = InternalPlayer.objects.filter(
                    register_number=reg_no,
                    booking__date__range=(week_start, week_end)
                ).count()

                override_allowed = RegisterOverride.objects.filter(
                    register_number=reg_no,
                    allow_more_than_two=True
                ).exists()

                if weekly_count >= 2 and not override_allowed:
                    raise ValueError(
                        "Contact admin – same register number should not repeat more than twice."
                    )

                players_data.append({
                    "name": person_name,
                    "reg_no": reg_no
                })

        except Exception as e:
            return render(request, "internal_details.html", {
                "error": str(e)
            })

        # ✅ SAVE DATA ATOMICALLY
        with transaction.atomic():
            internal_booking = InternalBooking.objects.create(
                name=name,
                email=request.session.get("student_email"),
                mobile=mobile,
                persons=persons,
                date=selected_date,
                slot=selected_slot,
                excel_file=excel_file
            )

            for p in players_data:
                InternalPlayer.objects.create(
                    booking=internal_booking,
                    player_name=p["name"],
                    register_number=p["reg_no"]
                )

        # ✅ Store for confirmation page
        request.session["internal_booking"] = {
            "customer_name": name,
            "mobile": mobile,
            "persons": persons,
            "players": players_data,
            "date": selected_date.strftime("%Y-%m-%d"),
            "slot": selected_slot
        }

        return redirect("booking:internal_confirmation")

    return render(request, "internal_details.html")


@never_cache
def internal_confirmation(request):
    if not request.session.get("internal_logged_in"):
        return redirect("booking:internal_login")

    data = request.session.get("internal_booking")

    if not data:
        return redirect("booking:internal_slots")

    return render(request, "internal_confirmation.html", data)


@never_cache
def internal_logout(request):
    request.session.flush()
    return redirect("booking:internal_login")


from django.core.mail import send_mail
from django.contrib.auth.hashers import make_password
from .models import PendingStudent, Student
import random


@never_cache
def signup(request):

    if request.method == "POST":

        username = request.POST.get("username", "").strip()
        email = request.POST.get("email", "").strip().lower()
        password = request.POST.get("password")
        confirm_password = request.POST.get("confirm_password")

        # Username validation
        if not username:
            return render(request, "signup.html", {
                "error": "Username is required."
            })

        # Allow only TCE Student Email IDs
        if not email.endswith("@student.tce.edu"):
            return render(request, "signup.html", {
                "error": "Please use your TCE Student Email ID (@student.tce.edu)."
            })

        # Password Match
        if password != confirm_password:
            return render(request, "signup.html", {
                "error": "Passwords do not match."
            })

        # Email already registered
        if Student.objects.filter(email=email).exists():
            return render(request, "signup.html", {
                "error": "This email is already registered."
            })

        # Generate 6-digit OTP
        otp = str(random.randint(100000, 999999))

        # Remove any previous pending signup
        PendingStudent.objects.filter(email=email).delete()

        # Save temporary signup details
        PendingStudent.objects.create(
            username=username,
            email=email,
            password=make_password(password),
            otp=otp
        )

        # Log OTP to Console (Dummy implementation)
        import logging
        logger = logging.getLogger(__name__)
        logger.warning("="*42)
        logger.warning(f"DUMMY OTP for {email} is: {otp}")
        logger.warning("="*42)
        
        # Fallback print with flush
        print(f"==========================================", flush=True)
        print(f"DUMMY OTP for {email} is: {otp}", flush=True)
        print(f"==========================================", flush=True)

        request.session["pending_email"] = email

        return redirect("booking:verify_otp")

    return render(request, "signup.html")


@never_cache
def verify_otp(request):

    email = request.session.get("pending_email")

    if not email:
        return redirect("booking:signup")

    pending = PendingStudent.objects.filter(email=email).first()

    if not pending:
        return redirect("booking:signup")

    if request.method == "POST":

        otp = request.POST.get("otp")

        if otp == pending.otp:

            Student.objects.create(
                username=pending.username,
                email=pending.email,
                password=pending.password
            )

            pending.delete()

            del request.session["pending_email"]

            return redirect("booking:internal_login")

        else:

            return render(
                request,
                "verify_otp.html",
                {
                    "error": "Invalid OTP"
                }
            )

    return render(request, "verify_otp.html")